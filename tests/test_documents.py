"""Document rendering: PDF, XLSX, CSV, JSON and HTML.

PDF tests are marked ``pdf`` and skip when the native stack is absent (a
Windows workstation, a slim image). CI installs pango/cairo and then asserts
the marked tests actually *ran* -- a silently skipping PDF suite is how a
renderer regression ships.
"""

from __future__ import annotations

import csv
import json
from datetime import date
from io import BytesIO, StringIO

import pytest
from flask import Flask
from flask.testing import FlaskClient

from massingbill.extensions import db
from massingbill.models import Role
from massingbill.services import application as app_service
from massingbill.services import sov as sov_service
from massingbill.services.money import cents
from massingbill.services.renderers import PDF_AVAILABLE, XLSX_AVAILABLE, available_formats
from massingbill.services.renderers import context as context_module
from massingbill.services.renderers.context import AIA_DISCLAIMER
from massingbill.services.renderers.documents import Format, render, render_package, store
from massingbill.services.renderers.tabular import G703_COLUMNS
from tests.factories import Tenant, make_tenant, sign_in

requires_pdf = pytest.mark.skipif(not PDF_AVAILABLE, reason="WeasyPrint native stack absent")

# The rendering extras are optional by design, and the `no-adapters` CI job
# installs the core only. Importing openpyxl at module scope made this whole
# file collapse there with ModuleNotFoundError -- which is the same graceful
# degradation the product promises, failing in the tests that assert it.
requires_xlsx = pytest.mark.skipif(not XLSX_AVAILABLE, reason="openpyxl not installed")

if XLSX_AVAILABLE:
    from openpyxl import load_workbook


@pytest.fixture
def tenant(app: Flask) -> Tenant:
    built = make_tenant("docs", contract_sum_cents=1_000_000_00)
    for item, description, value in (
        ("001", "Site work", 400_000_00),
        ("002", "Structure", 350_000_00),
        ("003", "Finishes", 250_000_00),
    ):
        sov_service.add_line(
            built.schedule,
            sov_service.LineInput(
                item_no=item, description=description, scheduled_value_cents=cents(value)
            ),
            actor=built.user(Role.OWNER),
        )
    sov_service.approve(built.schedule, actor=built.user(Role.OWNER))
    db.session.commit()
    return built


@pytest.fixture
def application(tenant: Tenant):
    built = app_service.open_period(
        tenant.contract,
        period_start=date(2026, 4, 1),
        period_end=date(2026, 4, 30),
        actor=tenant.user(Role.OWNER),
    )
    app_service.enter(
        built,
        [
            app_service.PeriodEntry(line_id=line.id, this_period=cents(work), stored=cents(stored))
            for line, (work, stored) in zip(
                built.lines, [(200_000_00, 0), (100_000_00, 25_000_00), (0, 0)], strict=True
            )
        ],
    )
    return built


# ── The view model ──────────────────────────────────────────────────────────


def test_the_view_carries_the_header_and_the_lines(application) -> None:
    view = context_module.build(application)

    assert view.number == 1
    assert view.header.line4 == 325_000_00
    assert len(view.lines) == 3
    assert view.lines[0].item_no == "001"
    assert view.lines[1].f == 25_000_00


def test_a_draft_renders_from_live_rows(application) -> None:
    assert context_module.build(application).from_snapshot is False


def test_a_submitted_application_renders_from_its_snapshot(tenant: Tenant, application) -> None:
    """The durability promise, made visible. After submission the document is
    built from the frozen snapshot, so a later schedule revision cannot restate
    what an owner already received."""
    app_service.submit(application, actor=tenant.user(Role.OWNER))
    db.session.commit()

    view = context_module.build(application)
    assert view.from_snapshot is True
    assert view.header.line8 == application.line8_current_payment_due


def test_the_view_includes_the_tieout_findings(application) -> None:
    view = context_module.build(application)

    assert view.tieout_ok
    assert view.findings
    assert any(f["rule_id"].startswith("INFO") for f in view.findings)


# ── CSV ─────────────────────────────────────────────────────────────────────


def test_csv_uses_the_g703_column_order(application) -> None:
    """The safe path: a licence holder populates their own official document
    from this, so the column order is part of the contract."""
    document = render(application, Format.CSV)
    rows = list(csv.reader(StringIO(document.content.decode("utf-8-sig"))))

    header = next(row for row in rows if row and row[0].startswith("A "))
    assert header == G703_COLUMNS


def test_csv_amounts_are_exact_decimal_strings(application) -> None:
    rows = list(csv.reader(StringIO(render(application, Format.CSV).content.decode("utf-8-sig"))))
    line = next(row for row in rows if row and row[0] == "001")

    assert line[3] == "400000.00"  # column C
    assert line[5] == "200000.00"  # column E


def test_csv_totals_row_sums_the_lines(application) -> None:
    rows = list(csv.reader(StringIO(render(application, Format.CSV).content.decode("utf-8-sig"))))
    totals = next(row for row in rows if len(row) > 1 and row[1] == "TOTALS")

    assert totals[3] == "1000000.00"
    assert totals[7] == "325000.00"  # column G


def test_csv_carries_the_disclaimer(application) -> None:
    body = render(application, Format.CSV).content.decode("utf-8-sig")
    assert "not affiliated with" in body


def test_csv_opens_cleanly_in_excel(application) -> None:
    """A BOM, or Excel mis-decodes anything non-ASCII in a description."""
    assert render(application, Format.CSV).content.startswith(b"\xef\xbb\xbf")


# ── JSON ────────────────────────────────────────────────────────────────────


def test_json_carries_every_g702_line(application) -> None:
    payload = json.loads(render(application, Format.JSON).content)

    for line in ("1", "2", "3", "4", "5a", "5b", "5", "6", "7", "8", "9"):
        assert f"line{line}" in payload["header"]

    assert payload["header"]["line4"]["cents"] == 325_000_00
    assert payload["header"]["line4"]["amount"] == "325000.00"


def test_json_lines_are_in_g703_order(application) -> None:
    payload = json.loads(render(application, Format.JSON).content)
    keys = list(payload["lines"][0])

    assert keys.index("c_scheduled_value_cents") < keys.index("d_previous_cents")
    assert keys.index("g_completed_stored_cents") < keys.index("h_balance_cents")


def test_json_includes_the_tieout_verdict(application) -> None:
    payload = json.loads(render(application, Format.JSON).content)

    assert payload["tieout"]["ok"] is True
    assert payload["tieout"]["findings"]
    assert payload["disclaimer"] == AIA_DISCLAIMER


# ── XLSX ────────────────────────────────────────────────────────────────────


def _workbook(application):
    return load_workbook(BytesIO(render(application, Format.XLSX).content))


@requires_xlsx
def test_the_workbook_has_the_three_sheets(application) -> None:
    assert _workbook(application).sheetnames == [
        "G702 Application",
        "G703 Continuation",
        "Reconciliation",
    ]


@requires_xlsx
def test_the_continuation_sheet_ships_live_formulas(application) -> None:
    """The most-requested export in the competitor corpus, and the most honest
    thing we produce: an owner's accountant can click a cell and see the
    arithmetic rather than being asked to trust a number."""
    sheet = _workbook(application)["G703 Continuation"]

    assert sheet["H6"].value == "=E6+F6+G6"  # column G = D + E + F
    assert sheet["J6"].value == "=D6-H6"  # column H = C - G
    assert sheet["I6"].value == "=IF(D6=0,0,H6/D6)"  # percent complete


@requires_xlsx
def test_the_workbook_totals_are_formulas_not_values(application) -> None:
    sheet = _workbook(application)["G703 Continuation"]
    totals_row = 6 + 3

    assert sheet[f"D{totals_row}"].value == f"=SUM(D6:D{totals_row - 1})"


@requires_xlsx
def test_the_cover_sheet_derives_from_the_continuation_sheet(application) -> None:
    sheet = _workbook(application)["G702 Application"]
    formulas = [
        cell.value
        for row in sheet.iter_rows(min_col=3, max_col=3)
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ]

    assert any("'G703 Continuation'!H" in f for f in formulas), "line 4 must reference the grid"
    assert any(f.count("-") == 1 and f.startswith("=C") for f in formulas), "lines 6/8/9 derive"


@requires_xlsx
def test_the_workbook_values_agree_with_the_engine(application) -> None:
    """The formulas must start from the right inputs, or they compute the wrong
    answer very transparently."""
    sheet = _workbook(application)["G703 Continuation"]

    assert float(sheet["D6"].value) == 400_000.00  # column C
    assert float(sheet["F6"].value) == 200_000.00  # column E
    assert float(sheet["G7"].value) == 25_000.00  # column F on line 002


@requires_xlsx
def test_the_workbook_carries_the_disclaimer(application) -> None:
    sheet = _workbook(application)["G703 Continuation"]
    text = " ".join(
        str(cell.value) for row in sheet.iter_rows() for cell in row if cell.value is not None
    )
    assert "not affiliated with" in text


@requires_xlsx
def test_the_reconciliation_sheet_lists_the_findings(application) -> None:
    sheet = _workbook(application)["Reconciliation"]
    rules = [sheet.cell(row=r, column=1).value for r in range(5, 25)]

    assert any(rule and rule.startswith("INFO") for rule in rules)


# ── HTML and PDF ────────────────────────────────────────────────────────────


def test_html_renders_the_whole_document(app: Flask, application) -> None:
    body = render(application, Format.HTML).content.decode("utf-8")

    assert "Application and Certificate for Payment" in body
    assert "Continuation Sheet" in body
    assert "Reconciliation" in body
    assert "$325,000.00" in body


def test_html_marks_a_draft_as_unsubmitted(app: Flask, application) -> None:
    """An unsubmitted document must not be mistakable for an issued one."""
    assert "DRAFT — NOT SUBMITTED" in render(application, Format.HTML).content.decode("utf-8")


def test_the_house_renderer_drops_the_aia_framing(app: Flask, application) -> None:
    application.form_style = "house"
    db.session.flush()

    body = render(application, Format.HTML).content.decode("utf-8")
    assert "Massing Bill standard form" in body
    assert "Application and Certificate for Payment" not in body


def test_the_disclaimer_cannot_be_removed_from_the_aia_style_renderer(
    app: Flask, application
) -> None:
    """docs/legal-forms-policy.md makes this binding, so it is asserted rather
    than trusted. Deleting this test is itself a policy violation."""
    body = render(application, Format.HTML).content.decode("utf-8")

    assert "not affiliated with, endorsed by, or sponsored by" in body
    assert "The American Institute of Architects" in body
    assert "G702" in body and "G703" in body


def test_even_the_house_renderer_carries_the_disclaimer(app: Flask, application) -> None:
    application.form_style = "house"
    db.session.flush()

    assert "not affiliated with" in render(application, Format.HTML).content.decode("utf-8")


@requires_pdf
@pytest.mark.pdf
def test_pdf_renders(app: Flask, application) -> None:
    document = render(application, Format.PDF)

    assert document.content.startswith(b"%PDF-")
    assert document.size > 2_000
    assert document.filename.endswith(".pdf")


@requires_pdf
@pytest.mark.pdf
def test_pdf_contains_the_numbers_and_the_disclaimer(app: Flask, application) -> None:
    from pypdf import PdfReader

    reader = PdfReader(BytesIO(render(application, Format.PDF).content))
    text = " ".join(page.extract_text() or "" for page in reader.pages)
    squashed = " ".join(text.split())

    assert "325,000.00" in squashed
    assert "Continuation Sheet" in squashed
    assert "not affiliated with" in squashed


@requires_pdf
@pytest.mark.pdf
def test_pdf_content_is_stable_across_renders(app: Flask, application) -> None:
    """The *content* is identical on every render.

    The raw bytes are not, and claiming otherwise was wrong: WeasyPrint stamps
    a creation timestamp into the PDF, so two renders a second apart differ by
    a few bytes. CI caught it.

    Byte-reproducibility is available to operators who want it, via
    ``SOURCE_DATE_EPOCH`` (asserted below) -- but it is not what proves a
    document is the one that was issued. The snapshot fingerprint does that,
    and it covers the numbers rather than the rendering.
    """
    from pypdf import PdfReader

    def text_of(content: bytes) -> str:
        reader = PdfReader(BytesIO(content))
        return " ".join(" ".join((page.extract_text() or "").split()) for page in reader.pages)

    assert text_of(render(application, Format.PDF).content) == text_of(
        render(application, Format.PDF).content
    )


@requires_pdf
@pytest.mark.pdf
def test_pdf_is_byte_reproducible_with_a_fixed_timestamp(
    app: Flask, application, monkeypatch: pytest.MonkeyPatch
) -> None:
    """WeasyPrint honours SOURCE_DATE_EPOCH, so a build that pins it gets
    identical bytes -- which is what an archival or evidentiary workflow wants."""
    monkeypatch.setenv("SOURCE_DATE_EPOCH", "1767225600")  # 2026-01-01T00:00:00Z

    assert render(application, Format.PDF).content == render(application, Format.PDF).content


def test_pdf_reports_why_it_is_unavailable_rather_than_crashing(app: Flask, application) -> None:
    if PDF_AVAILABLE:
        pytest.skip("PDF rendering is available here")

    from massingbill.errors import AdapterUnavailableError

    with pytest.raises(AdapterUnavailableError, match="WeasyPrint"):
        render(application, Format.PDF)


# ── Packaging and storage ───────────────────────────────────────────────────


def test_a_package_contains_every_available_format(app: Flask, application) -> None:
    package = render_package(application)

    assert {d.fmt for d in package.documents} == set(available_formats())
    assert all(d.content for d in package.documents)


def test_documents_are_identified_by_digest(app: Flask, application) -> None:
    document = render(application, Format.CSV)

    assert len(document.sha256) == 64
    assert render(application, Format.CSV).sha256 == document.sha256


def test_filenames_carry_the_project_and_application_number(app: Flask, application) -> None:
    assert render(application, Format.CSV).filename == "2026-001-pay-app-001.csv"


def test_storing_a_document_returns_a_pointer(app: Flask, application) -> None:
    document = render(application, Format.JSON)
    pointer = store(application, document)

    assert pointer.backend == "local"
    assert pointer.sha256 == document.sha256
    assert application.organization_id in pointer.key


# ── The download route ──────────────────────────────────────────────────────


def test_downloading_json(client: FlaskClient, tenant: Tenant, application) -> None:
    sign_in(client, tenant.user(Role.OWNER))
    url = f"/projects/{tenant.project.id}/applications/{application.id}/download.json"
    response = client.get(url)

    assert response.status_code == 200
    assert response.headers["Content-Type"].startswith("application/json")
    assert "attachment" in response.headers["Content-Disposition"]
    assert len(response.headers["X-Document-SHA256"]) == 64


def test_downloading_csv(client: FlaskClient, tenant: Tenant, application) -> None:
    sign_in(client, tenant.user(Role.OWNER))
    base = f"/projects/{tenant.project.id}/applications/{application.id}/download"

    assert client.get(f"{base}.csv").status_code == 200


@requires_xlsx
def test_downloading_xlsx(client: FlaskClient, tenant: Tenant, application) -> None:
    sign_in(client, tenant.user(Role.OWNER))
    base = f"/projects/{tenant.project.id}/applications/{application.id}/download"

    assert client.get(f"{base}.xlsx").status_code == 200


def test_an_unknown_format_is_not_found(client: FlaskClient, tenant: Tenant, application) -> None:
    sign_in(client, tenant.user(Role.OWNER))
    url = f"/projects/{tenant.project.id}/applications/{application.id}/download.docx"

    assert client.get(url).status_code == 404


def test_an_unavailable_format_explains_itself(
    client: FlaskClient, tenant: Tenant, application
) -> None:
    if PDF_AVAILABLE:
        pytest.skip("PDF rendering is available here")

    sign_in(client, tenant.user(Role.OWNER))
    url = f"/projects/{tenant.project.id}/applications/{application.id}/download.pdf"
    response = client.get(url)

    assert response.status_code == 503
    assert b"WeasyPrint" in response.data


def test_downloads_are_tenant_scoped(
    client: FlaskClient, tenant: Tenant, application, app: Flask
) -> None:
    other = make_tenant("rival")
    sign_in(client, other.user(Role.OWNER))

    url = f"/projects/{tenant.project.id}/applications/{application.id}/download.json"
    assert client.get(url).status_code == 404


def test_a_viewer_may_download(client: FlaskClient, tenant: Tenant, application) -> None:
    sign_in(client, tenant.user(Role.VIEWER))
    url = f"/projects/{tenant.project.id}/applications/{application.id}/download.json"

    assert client.get(url).status_code == 200

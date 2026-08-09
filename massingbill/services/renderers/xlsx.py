"""The spreadsheet renderer.

**The G703 sheet ships live formulas, not values.** ``=D6+E6+F6``, ``=C6-G6``,
``=G6/C6`` -- and the G702 sheet references the G703 totals rather than
restating them. An owner's accountant can click any cell and see the
arithmetic.

That is the single most-requested export in the competitor review corpus, and
it is also the most honest thing we produce: a spreadsheet of frozen numbers
asks to be trusted, a spreadsheet of formulas can be checked.

Column order matches G703 exactly (docs/legal-forms-policy.md), so a customer
holding an AIA licence can populate their own official document from it.
"""

from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from typing import cast

from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from massingbill.services.money import cents, to_decimal
from massingbill.services.renderers.context import ApplicationView

MONEY_FORMAT = "#,##0.00;[Red]-#,##0.00"
PERCENT_FORMAT = "0.00%"

_HEADING = Font(bold=True, size=11)
_LABEL = Font(size=10)
_TOTAL = Font(bold=True, size=10)
_MUTED = Font(size=8, italic=True, color="666666")
_FILL = PatternFill("solid", fgColor="F2F4F7")
_THIN = Side(style="thin", color="D0D5DD")
_BOX = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def render(view: ApplicationView) -> bytes:
    """Build the workbook and return it as bytes."""
    book = Workbook()

    continuation = cast(Worksheet, book.active)
    continuation.title = "G703 Continuation"
    _continuation_sheet(continuation, view)

    cover = book.create_sheet("G702 Application", 0)
    _cover_sheet(cover, view, continuation.title)

    _tieout_sheet(book.create_sheet("Reconciliation"), view)

    buffer = BytesIO()
    book.save(buffer)
    return buffer.getvalue()


# ── G703 ────────────────────────────────────────────────────────────────────

#: Row the line items start on. Everything below is computed from it, so the
#: layout can move without the formulas silently pointing at the wrong cells.
_FIRST_LINE_ROW = 6


def _continuation_sheet(sheet: Worksheet, view: ApplicationView) -> None:
    sheet["A1"] = "Continuation Sheet"
    sheet["A1"].font = _HEADING
    sheet["A2"] = f"{view.project_number} — {view.project_name}"
    sheet["A3"] = f"Application No. {view.number} · {view.period_label}"
    for ref in ("A2", "A3"):
        sheet[ref].font = _LABEL

    headings = [
        ("A", "A\nItem No."),
        ("B", "B\nDescription of Work"),
        ("C", "CSI"),
        ("D", "C\nScheduled Value"),
        ("E", "D\nFrom Previous Application"),
        ("F", "E\nThis Period"),
        ("G", "F\nMaterials Presently Stored"),
        ("H", "G\nTotal Completed and Stored"),
        ("I", "%\n(G / C)"),
        ("J", "H\nBalance to Finish"),
        ("K", "I\nRetainage"),
    ]
    for column, title in headings:
        cell = sheet[f"{column}5"]
        cell.value = title
        cell.font = _HEADING
        cell.fill = _FILL
        cell.border = _BOX
        cell.alignment = Alignment(wrap_text=True, vertical="bottom")

    row = _FIRST_LINE_ROW
    for line in view.lines:
        sheet[f"A{row}"] = line.item_no
        sheet[f"B{row}"] = line.description
        sheet[f"C{row}"] = line.csi_code
        sheet[f"D{row}"] = to_decimal(cents(line.c))
        sheet[f"E{row}"] = to_decimal(cents(line.d))
        sheet[f"F{row}"] = to_decimal(cents(line.e))
        sheet[f"G{row}"] = to_decimal(cents(line.f))

        # The point of the whole sheet: G, %, H are derived in the workbook, so
        # anyone can see how they were reached rather than taking them on trust.
        sheet[f"H{row}"] = f"=E{row}+F{row}+G{row}"
        sheet[f"I{row}"] = f"=IF(D{row}=0,0,H{row}/D{row})"
        sheet[f"J{row}"] = f"=D{row}-H{row}"
        sheet[f"K{row}"] = to_decimal(cents(line.i))

        for column in "DEFGHJK":
            sheet[f"{column}{row}"].number_format = MONEY_FORMAT
        sheet[f"I{row}"].number_format = PERCENT_FORMAT
        for column in "ABCDEFGHIJK":
            sheet[f"{column}{row}"].border = _BOX
        row += 1

    last = row - 1
    sheet[f"B{row}"] = "TOTALS"
    sheet[f"B{row}"].font = _TOTAL
    for column in "DEFGHJK":
        cell = sheet[f"{column}{row}"]
        cell.value = f"=SUM({column}{_FIRST_LINE_ROW}:{column}{last})"
        cell.number_format = MONEY_FORMAT
        cell.font = _TOTAL
        cell.border = _BOX

    sheet.freeze_panes = f"A{_FIRST_LINE_ROW}"
    widths = {
        "A": 10,
        "B": 42,
        "C": 8,
        "D": 16,
        "E": 16,
        "F": 14,
        "G": 16,
        "H": 18,
        "I": 9,
        "J": 16,
        "K": 14,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    note = row + 2
    sheet[f"A{note}"] = view.disclaimer
    sheet[f"A{note}"].font = _MUTED
    sheet[f"A{note}"].alignment = Alignment(wrap_text=True, vertical="top")
    sheet.merge_cells(start_row=note, start_column=1, end_row=note + 2, end_column=11)


# ── G702 ────────────────────────────────────────────────────────────────────


def _cover_sheet(sheet: Worksheet, view: ApplicationView, continuation: str) -> None:
    sheet["A1"] = "Application and Certificate for Payment"
    sheet["A1"].font = Font(bold=True, size=13)

    facts: list[tuple[str, str]] = [
        ("Project", f"{view.project_number} — {view.project_name}"),
        ("Contract", view.contract_number),
        ("Application No.", str(view.number)),
        ("Period", view.period_label),
        ("Application date", view.application_date.isoformat()),
    ]
    row = 3
    for label, value in facts:
        sheet[f"A{row}"] = label
        sheet[f"A{row}"].font = _LABEL
        sheet[f"C{row}"] = value
        row += 1

    row += 1
    totals = f"'{continuation}'!"
    line_row = row

    # Lines 4 and 5 reference the continuation sheet, and 3/6/8/9 are computed
    # from the lines above them -- so the cover sheet demonstrably agrees with
    # the grid instead of asserting it separately.
    entries: list[tuple[str, str, Decimal | str | None]] = [
        ("1", "Original contract sum", to_decimal(cents(view.header.line1))),
        ("2", "Net change by change orders", to_decimal(cents(view.header.line2))),
        ("3", "Contract sum to date (1 + 2)", f"=C{line_row}+C{line_row + 1}"),
        ("4", "Total completed and stored to date", None),  # filled below
        ("5a", "Retainage on completed work", to_decimal(cents(view.header.line5a))),
        ("5b", "Retainage on stored material", to_decimal(cents(view.header.line5b))),
        ("5", "Total retainage (5a + 5b)", f"=C{line_row + 4}+C{line_row + 5}"),
        ("6", "Total earned less retainage (4 − 5)", f"=C{line_row + 3}-C{line_row + 6}"),
        ("7", "Less previous certificates for payment", to_decimal(cents(view.header.line7))),
        ("8", "Current payment due (6 − 7)", f"=C{line_row + 7}-C{line_row + 8}"),
        (
            "9",
            "Balance to finish, including retainage (3 − 6)",
            f"=C{line_row + 2}-C{line_row + 7}",
        ),
    ]

    for number, line_label, line_value in entries:
        sheet[f"A{row}"] = number
        sheet[f"B{row}"] = line_label
        cell = sheet[f"C{row}"]
        cell.value = line_value
        cell.number_format = MONEY_FORMAT
        if number in {"3", "6", "8", "9"}:
            sheet[f"A{row}"].font = _TOTAL
            sheet[f"B{row}"].font = _TOTAL
            cell.font = _TOTAL
        row += 1

    # Line 4 comes straight from the continuation sheet's column G total.
    sheet[f"C{line_row + 3}"] = f"={totals}H{_continuation_totals_row(view)}"
    sheet[f"C{line_row + 3}"].number_format = MONEY_FORMAT
    sheet[f"C{line_row + 3}"].font = _TOTAL

    row += 1
    sheet[f"A{row}"] = "Change order summary"
    sheet[f"A{row}"].font = _HEADING
    row += 1
    for co_label, co_amount in (
        ("Approved in previous months — additions", view.header.co_prev_additions),
        ("Approved in previous months — deductions", view.header.co_prev_deductions),
        ("Approved this month — additions", view.header.co_this_additions),
        ("Approved this month — deductions", view.header.co_this_deductions),
    ):
        sheet[f"B{row}"] = co_label
        sheet[f"C{row}"] = to_decimal(cents(co_amount))
        sheet[f"C{row}"].number_format = MONEY_FORMAT
        row += 1

    row += 2
    sheet[f"A{row}"] = view.disclaimer
    sheet[f"A{row}"].font = _MUTED
    sheet[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")
    sheet.merge_cells(start_row=row, start_column=1, end_row=row + 2, end_column=4)

    sheet.column_dimensions["A"].width = 6
    sheet.column_dimensions["B"].width = 46
    sheet.column_dimensions["C"].width = 20


def _continuation_totals_row(view: ApplicationView) -> int:
    return _FIRST_LINE_ROW + len(view.lines)


# ── Reconciliation ──────────────────────────────────────────────────────────


def _tieout_sheet(sheet: Worksheet, view: ApplicationView) -> None:
    sheet["A1"] = "Reconciliation"
    sheet["A1"].font = _HEADING
    sheet["A2"] = view.tieout_summary
    sheet["A2"].font = _LABEL

    for index, title in enumerate(
        ["Rule", "Severity", "Finding", "Expected", "Actual", "Difference", "Reference"], start=1
    ):
        cell = sheet.cell(row=4, column=index, value=title)
        cell.font = _HEADING
        cell.fill = _FILL
        cell.border = _BOX

    row = 5
    for finding in view.findings:
        sheet.cell(row=row, column=1, value=finding["rule_id"])
        sheet.cell(row=row, column=2, value=finding["severity"])
        sheet.cell(row=row, column=3, value=finding["message"])
        for offset, key in enumerate(("expected_cents", "actual_cents", "delta_cents"), start=4):
            value = finding.get(key)
            cell = cast(Cell, sheet.cell(row=row, column=offset))
            if value is not None:
                cell.value = to_decimal(cents(int(value)))
                cell.number_format = MONEY_FORMAT
        sheet.cell(row=row, column=7, value=finding.get("citation", ""))
        row += 1

    for column, width in zip("ABCDEFG", (14, 11, 70, 16, 16, 14, 44), strict=True):
        sheet.column_dimensions[column].width = width
    sheet.column_dimensions[get_column_letter(3)].width = 70
